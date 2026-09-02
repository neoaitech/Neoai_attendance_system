import os
import pytest
from datetime import date
import openpyxl
from backend.app.services.report_service import report_service
from backend.app.db.models import AttendanceSession, AttendanceRecord, ClassCourse, Student

@pytest.fixture
def report_class(db_session, teacher_user, sample_students):
    course = ClassCourse(
        code=f"REP-{os.urandom(4).hex()}",
        name="Reporting Course",
        department="Computer Science",
        semester="Fall 2026",
        section="A",
        teacher_id=teacher_user.id
    )
    for s in sample_students:
        course.students.append(s)
    db_session.add(course)
    db_session.commit()
    db_session.refresh(course)
    return course

def test_report_service_class_summary(db_session, report_class, sample_students, teacher_user):
    # Create 2 sessions specifically for report_class
    for i in range(1, 3):
        sess = AttendanceSession(
            class_id=report_class.id,
            teacher_id=teacher_user.id,
            session_name=f"Report Test Session {i}",
            session_date=date.today(),
            status="CONFIRMED"
        )
        db_session.add(sess)
        db_session.flush()

        # Student 0 present both (100%), Student 1 present one (50% defaulter), Student 2 absent both (0% defaulter)
        db_session.add(AttendanceRecord(session_id=sess.id, student_id=sample_students[0].id, status="PRESENT"))
        db_session.add(AttendanceRecord(session_id=sess.id, student_id=sample_students[1].id, status="PRESENT" if i == 1 else "ABSENT"))
        db_session.add(AttendanceRecord(session_id=sess.id, student_id=sample_students[2].id, status="ABSENT"))

    db_session.commit()

    summary = report_service.get_class_summary_data(db_session, report_class.id)
    assert summary["class_id"] == report_class.id
    assert summary["total_sessions_conducted"] == 2
    assert summary["total_enrolled"] == 3

    # Student 0: 100%
    s0 = next(s for s in summary["students_summary"] if s["student_id"] == sample_students[0].id)
    assert s0["attendance_percentage"] == 100.0
    assert s0["is_defaulter"] is False

    # Student 1: 50%
    s1 = next(s for s in summary["students_summary"] if s["student_id"] == sample_students[1].id)
    assert s1["attendance_percentage"] == 50.0
    assert s1["is_defaulter"] is True

    # Student 2: 0%
    s2 = next(s for s in summary["students_summary"] if s["student_id"] == sample_students[2].id)
    assert s2["attendance_percentage"] == 0.0
    assert s2["is_defaulter"] is True

    assert summary["defaulters_count"] == 2

def test_excel_export_generation(db_session, report_class, sample_students, teacher_user):
    filepath = report_service.export_excel(db_session, report_class.id)
    assert os.path.exists(filepath)
    assert filepath.endswith(".xlsx")

    # Load with openpyxl to verify valid multi-sheet workbook structure
    wb = openpyxl.load_workbook(filepath)
    assert "Executive Summary" in wb.sheetnames
    assert any("Div" in s for s in wb.sheetnames)
    ws = wb["Executive Summary"]
    assert "VisionAttend Pro" in ws["A1"].value
    wb.close()

def test_advanced_multi_sheet_export(db_session, report_class, sample_students, teacher_user):
    # Test date-range and multi-sheet generation
    filepath = report_service.export_advanced_excel(
        db=db_session,
        class_id=report_class.id,
        division="ALL",
        start_date=date.today(),
        end_date=date.today()
    )
    assert os.path.exists(filepath)
    wb = openpyxl.load_workbook(filepath)
    assert "Executive Summary" in wb.sheetnames
    assert len(wb.sheetnames) >= 2
    wb.close()

def test_pdf_export_generation(db_session, report_class, sample_students, teacher_user):
    filepath = report_service.export_pdf(db_session, report_class.id)
    assert os.path.exists(filepath)
    assert filepath.endswith(".pdf")
    assert os.path.getsize(filepath) > 1000

def test_multi_select_reports_and_export(db_session, report_class, sample_students, teacher_user):
    # Test multi-program, multi-semester, and multi-division filter combination
    data = report_service.get_advanced_report_data(
        db=db_session,
        department="Computer Science",
        programs=["B.Tech", "MCA"],
        semesters=["Semester 5", "Fall 2026"],
        divisions=["A", "B"]
    )
    assert data is not None
    assert "divisions" in data
    assert "Executive" in data["scope_title"] or "Computer Science" in data["scope_title"]
    assert "total_enrolled" in data

    # Test Excel Export with multi-select parameters
    excel_path = report_service.export_advanced_excel(
        db=db_session,
        department="Computer Science",
        programs="B.Tech,MCA",
        semesters="Semester 5,Fall 2026",
        divisions="A,B"
    )
    assert os.path.exists(excel_path)
    wb = openpyxl.load_workbook(excel_path)
    assert "Executive Summary" in wb.sheetnames
    wb.close()

    # Test PDF Export with multi-select parameters
    pdf_path = report_service.export_advanced_pdf(
        db=db_session,
        department="Computer Science",
        programs="B.Tech,MCA",
        semesters="Semester 5,Fall 2026",
        divisions="A,B"
    )
    assert os.path.exists(pdf_path)
    assert os.path.getsize(pdf_path) > 1000

def test_individual_student_report_and_bunk_log(db_session, report_class, sample_students, teacher_user):
    # Student 1 has 1 Present and 1 Absent (Bunk)
    student = sample_students[1]
    data = report_service.get_student_detailed_report(db_session, student.id)
    assert data is not None
    assert data["student_id"] == student.id
    assert data["roll_number"] == student.roll_number
    assert data["total_lectures_conducted"] >= 2
    assert "subjects_breakdown" in data
    assert len(data["subjects_breakdown"]) >= 1
    assert "lecture_history" in data
    assert any(h["is_bunk"] is True for h in data["lecture_history"])

    # Test individual student PDF export
    pdf_path = report_service.export_student_pdf(db_session, student.id)
    assert os.path.exists(pdf_path)
    assert pdf_path.endswith(".pdf")
    assert os.path.getsize(pdf_path) > 1000


