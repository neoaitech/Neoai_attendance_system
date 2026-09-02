import os
import uuid
import json
from datetime import datetime, date, time
from pathlib import Path
from typing import Dict, Any, List, Optional, Union
from sqlalchemy.orm import Session
from sqlalchemy import func
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, A4, landscape, portrait
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, KeepTogether, HRFlowable, PageBreak, Image as RLImage
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from backend.app.core.config import settings
from backend.app.core.datetime_utils import format_ist_time, format_iso_utc, format_ist_date
from backend.app.db.models import (
    ClassCourse, Student, AttendanceSession, AttendanceRecord, User
)

def _parse_multi(val: Any) -> List[str]:
    """
    Safely parses single string, comma-separated string, or list into cleaned unique list.
    Filters out 'ALL', empty strings, and None.
    """
    if val is None:
        return []
    if isinstance(val, list):
        res = []
        for v in val:
            if isinstance(v, str):
                for part in v.split(","):
                    p = part.strip()
                    if p and p.upper() != "ALL" and p not in res:
                        res.append(p)
            elif v is not None and str(v).strip() not in res:
                res.append(str(v).strip())
        return res
    if isinstance(val, str):
        if val.strip().upper() == "ALL" or not val.strip():
            return []
        parts = [p.strip() for p in val.split(",") if p.strip() and p.strip().upper() != "ALL"]
        return list(dict.fromkeys(parts))
    return [str(val).strip()]


class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            super().showPage()
        super().save()

    def draw_page_decorations(self, page_count):
        self.saveState()
        self.setFont("Helvetica-Bold", 7)
        self.setFillColor(colors.HexColor("#475569"))
        
        # Header (Top line)
        self.drawString(28, 818, "VISIONATTEND PRO — STUDENT ATTENDANCE & ACADEMIC AUDIT DOSSIER")
        self.setFont("Helvetica", 7)
        self.drawRightString(595 - 28, 818, "Official Institutional Record")
        self.setStrokeColor(colors.HexColor("#CBD5E1"))
        self.setLineWidth(0.5)
        self.line(28, 812, 595 - 28, 812)
        
        # Footer (Bottom line)
        self.line(28, 36, 595 - 28, 36)
        now_str = datetime.now().strftime("%d-%b-%Y %I:%M %p")
        self.drawString(28, 24, f"Generated: {now_str} • Confidential Institutional Academic Audit")
        self.drawRightString(595 - 28, 24, f"Page {self._pageNumber} of {page_count}")
        
        self.restoreState()


class ReportService:
    @staticmethod
    def get_available_filters(db: Session) -> Dict[str, Any]:
        """
        Returns all unique filter options for courses, programs, semesters, departments,
        divisions, and available attendance dates.
        """
        courses = db.query(ClassCourse).all()
        students = db.query(Student).all()
        sessions = db.query(AttendanceSession).all()

        default_programs = ["BCA", "MCA", "MBA", "BBA", "BA", "MA", "B.Tech", "M.Tech"]
        existing_programs = {getattr(c, "program", None) for c in courses if getattr(c, "program", None)} | {getattr(s, "program", None) for s in students if getattr(s, "program", None)}
        # Preserve preferred order of standard programs, then append any custom ones
        programs = [p for p in default_programs if p in existing_programs or True] + [p for p in sorted(list(existing_programs)) if p and p not in default_programs]

        semesters = sorted(list({c.semester for c in courses if c.semester} | {s.semester for s in students if s.semester}))
        if not semesters:
            semesters = [f"Semester {i}" for i in range(1, 9)]

        departments = sorted(list({c.department for c in courses if c.department} | {s.department for s in students if s.department}))
        if not departments:
            departments = [
                "Computer Science & Engineering",
                "Artificial Intelligence & Data Science",
                "Information Technology",
                "Electronics & Telecommunication",
                "Mechanical Engineering",
                "Civil Engineering"
            ]
        
        divisions = sorted(list({c.section for c in courses if c.section} | {s.section for s in students if s.section}))
        if not divisions:
            divisions = ["A", "B", "C", "D"]

        session_dates = [s.session_date for s in sessions if s.session_date]
        min_date = min(session_dates).isoformat() if session_dates else date.today().replace(day=1).isoformat()
        max_date = max(session_dates).isoformat() if session_dates else date.today().isoformat()

        return {
            "courses": [
                {
                    "id": c.id,
                    "code": c.code,
                    "name": c.name,
                    "program": getattr(c, "program", "B.Tech"),
                    "semester": c.semester,
                    "department": c.department,
                    "section": c.section
                } for c in courses
            ],
            "programs": programs,
            "semesters": semesters,
            "departments": departments,
            "divisions": divisions,
            "date_range": {
                "min_date": min_date,
                "max_date": max_date,
                "today": date.today().isoformat()
            }
        }

    @staticmethod
    def get_advanced_report_data(
        db: Session,
        class_id: Optional[int] = None,
        class_ids: Optional[Any] = None,
        department: Optional[str] = None,
        departments: Optional[Any] = None,
        program: Optional[str] = None,
        programs: Optional[Any] = None,
        semester: Optional[str] = None,
        semesters: Optional[Any] = None,
        division: Optional[str] = None,
        divisions: Optional[Any] = None,
        attendance_type: Optional[str] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> Dict[str, Any]:
        """
        Gathers comprehensive attendance data organized hierarchically by:
        Program (e.g. BCA, MCA, B.Tech) -> Semester (e.g. Sem 1, Sem 3) -> Division (e.g. Div A, Div B).
        Accurately calculates Normal Course Attendance and Extra Lecture Attendance.
        """
        dept_val = department or departments
        programs_list = _parse_multi(programs or program)
        semesters_list = _parse_multi(semesters or semester)
        divisions_list = _parse_multi(divisions or division)
        att_type_filter = (attendance_type or "ALL").strip().upper()

        c_ids_list = []
        if class_id:
            c_ids_list.append(class_id)
        if class_ids:
            for x in _parse_multi(class_ids):
                try:
                    c_ids_list.append(int(x))
                except ValueError:
                    pass

        # 1. Determine Target Courses
        course_query = db.query(ClassCourse)
        if c_ids_list:
            course_query = course_query.filter(ClassCourse.id.in_(c_ids_list))
        if dept_val and dept_val.upper() != "ALL":
            course_query = course_query.filter(ClassCourse.department == dept_val)
        if programs_list:
            course_query = course_query.filter(ClassCourse.program.in_(programs_list))
        if semesters_list:
            course_query = course_query.filter(ClassCourse.semester.in_(semesters_list))
        if divisions_list:
            course_query = course_query.filter(ClassCourse.section.in_(divisions_list))
        
        target_courses = course_query.all()
        target_course_ids = [c.id for c in target_courses]

        # 2. Fetch Matching Attendance Sessions in Date Range
        session_query = db.query(AttendanceSession)
        has_specific_filter = bool(c_ids_list or (dept_val and dept_val.upper() != "ALL") or programs_list or semesters_list or divisions_list)
        if target_course_ids:
            session_query = session_query.filter(AttendanceSession.class_id.in_(target_course_ids))
        elif has_specific_filter:
            session_query = session_query.filter(AttendanceSession.id == -1)

        if start_date:
            session_query = session_query.filter(AttendanceSession.session_date >= start_date)
        if end_date:
            session_query = session_query.filter(AttendanceSession.session_date <= end_date)

        sessions = session_query.order_by(AttendanceSession.session_date.asc(), AttendanceSession.id.asc()).all()
        session_ids = [s.id for s in sessions]

        unique_session_dates = sorted(list({s.session_date for s in sessions if s.session_date}))
        unique_date_strs = [d.isoformat() for d in unique_session_dates]

        # 3. Identify Students in Selected Scope
        students_query = db.query(Student).filter(Student.is_active == True)
        
        if c_ids_list and len(c_ids_list) == 1 and target_courses:
            students = list(target_courses[0].students)
        else:
            if dept_val and dept_val.upper() != "ALL":
                students_query = students_query.filter(Student.department == dept_val)
            if programs_list:
                students_query = students_query.filter(Student.program.in_(programs_list))
            if semesters_list:
                students_query = students_query.filter(Student.semester.in_(semesters_list))
            if divisions_list:
                students_query = students_query.filter(Student.section.in_(divisions_list))
            students = students_query.all()

        # 4. Group Students Hierarchically: Program -> Semester -> Division
        batch_map: Dict[str, List[Student]] = {}
        for s in students:
            prog = getattr(s, "program", None) or "B.Tech"
            sem = getattr(s, "semester", None) or "Semester 7"
            sec = (s.section or "A").strip().upper()

            if programs_list and prog.upper() not in [p.upper() for p in programs_list]:
                continue
            if semesters_list and sem.upper() not in [sm.upper() for sm in semesters_list]:
                continue
            if divisions_list and sec not in [d.upper() for d in divisions_list]:
                continue

            batch_key = f"{prog}___{sem}___{sec}"
            if batch_key not in batch_map:
                batch_map[batch_key] = []
            batch_map[batch_key].append(s)

        if not batch_map and not target_courses:
            batch_map["B.Tech___Semester 7___A"] = []

        total_sessions_count = len(sessions)
        overall_defaulters = 0
        total_students_count = 0
        total_extra_lectures_conducted = 0

        batches_list: List[Dict[str, Any]] = []
        hierarchy: Dict[str, Dict[str, Dict[str, Any]]] = {}
        divisions_data: Dict[str, Any] = {}

        for b_key in sorted(batch_map.keys()):
            parts = b_key.split("___")
            prog, sem, sec = parts[0], parts[1], parts[2]
            b_students = sorted(batch_map[b_key], key=lambda x: (x.roll_number or ""))
            
            b_students_summary = []
            b_pct_sum = 0.0
            b_defaulters = 0
            b_extra_total = 0

            # Match regular course sessions for this specific batch strictly (case-insensitive with section fallback)
            if c_ids_list:
                b_course_ids = [c.id for c in target_courses]
            else:
                b_course_ids = [
                    c.id for c in target_courses 
                    if (getattr(c, "program", None) or "B.Tech").strip().upper() == prog.strip().upper() 
                    and (c.semester or "Semester 7").strip().upper() == sem.strip().upper() 
                    and ((c.section or "A").strip().upper() == sec.strip().upper() or (c.section or "").strip().upper() in ["ALL", "BOTH", "*", "COMBINED"])
                ]
            b_sessions = [sess for sess in sessions if sess.class_id in b_course_ids] if b_course_ids else []
            b_session_ids = [sess.id for sess in b_sessions]
            b_sessions_count = len(b_sessions)

            for s in b_students:
                # Determine student-specific enrolled curriculum courses if assigned
                s_prog = (getattr(s, "program", None) or "").strip().upper()
                s_enrolled = [
                    c for c in s.enrolled_classes
                    if not s_prog 
                    or not getattr(c, "program", None) 
                    or c.program.strip().upper() in ["ALL", "*", "ANY", s_prog]
                ] if getattr(s, "enrolled_classes", None) and len(s.enrolled_classes) > 0 else []

                s_course_ids = [c.id for c in s_enrolled] if s_enrolled else b_course_ids
                s_sessions = [sess for sess in sessions if sess.class_id in s_course_ids] if s_course_ids else []
                s_session_ids = [sess.id for sess in s_sessions]
                s_sessions_count = len(s_sessions)

                # 1. Regular Attendance Records for enrolled course sessions
                reg_records = []
                if s_session_ids:
                    reg_records = db.query(AttendanceRecord).filter(
                        AttendanceRecord.session_id.in_(s_session_ids),
                        AttendanceRecord.student_id == s.id,
                        AttendanceRecord.is_extra_lecture == False,
                        AttendanceRecord.verification_type != "EXTRA_LECTURE",
                        AttendanceRecord.attendance_type != "EXTRA_LECTURE"
                    ).all()

                # 2. Extra Lecture Attendance Records (Student attended outside classes or flagged Extra Lecture, deduplicated by session)
                extra_records_query = db.query(AttendanceRecord).join(AttendanceSession, AttendanceRecord.session_id == AttendanceSession.id).filter(
                    AttendanceRecord.student_id == s.id,
                    AttendanceRecord.status.in_(["PRESENT", "LATE"]),
                    (AttendanceRecord.is_extra_lecture == True) | 
                    (AttendanceRecord.verification_type == "EXTRA_LECTURE") | 
                    (AttendanceRecord.attendance_type == "EXTRA_LECTURE") |
                    (~AttendanceRecord.session_id.in_(s_session_ids) if s_session_ids else True)
                )
                if start_date:
                    extra_records_query = extra_records_query.filter(AttendanceSession.session_date >= start_date)
                if end_date:
                    extra_records_query = extra_records_query.filter(AttendanceSession.session_date <= end_date)
                
                extra_records = extra_records_query.all()
                distinct_extra_session_ids = set(er.session_id for er in extra_records if er.session_id not in s_session_ids or er.is_extra_lecture)
                s_extra_count = len(distinct_extra_session_ids)
                b_extra_total += s_extra_count

                # Build Daily Status Matrix (attended/total per day)
                session_status_map: Dict[int, str] = {r.session_id: r.status for r in reg_records}
                daily_map: Dict[str, Any] = {}
                matrix_sessions = s_sessions if s_enrolled else b_sessions
                for sess in matrix_sessions:
                    if sess.session_date:
                        d_str = sess.session_date.isoformat()
                        if d_str not in daily_map:
                            daily_map[d_str] = {"attended": 0, "total": 0}
                        daily_map[d_str]["total"] += 1
                        st = session_status_map.get(sess.id, "A")
                        if st in ["PRESENT", "LATE"]:
                            daily_map[d_str]["attended"] += 1

                # Also include extra lecture sessions attended by this student on each date
                seen_extra_sess_ids = set()
                for er in extra_records:
                    sess = er.session
                    if sess and sess.id in distinct_extra_session_ids and sess.id not in seen_extra_sess_ids:
                        seen_extra_sess_ids.add(sess.id)
                        if sess.session_date:
                            d_str = sess.session_date.isoformat()
                            if d_str not in daily_map:
                                daily_map[d_str] = {"attended": 0, "total": 0}
                            daily_map[d_str]["attended"] += 1
                            daily_map[d_str]["total"] += 1

                frozen_count = sum(1 for r in reg_records if r.status == "FROZEN")
                present_count = sum(1 for r in reg_records if r.status in ["PRESENT", "LATE"])
                cur_total_sess = max(0, s_sessions_count - frozen_count)
                absent_count = max(0, cur_total_sess - present_count)
                normal_pct = round((present_count / cur_total_sess) * 100.0, 2) if cur_total_sess > 0 else 0.0

                # EXACT REQUIRED COMBINED LOGIC (Exempting Frozen Sessions):
                # Normal Sessions = regular course lectures (exempting frozen sessions)
                # Normal Present = student attended regular course lectures
                # Normal Absent = regular sessions not attended
                # Extra Lectures = approved extra-curricular / cross-division lectures
                # Total Sessions = Normal Sessions + Extra Lectures
                # Total Present = Normal Present + Extra Lectures
                # Total Absent = Total Sessions - Total Present (= Normal Absent)
                # Final Attendance % = (Total Present / Total Sessions) * 100
                # Defaulter Status = Final Attendance % < threshold
                total_student_sessions = cur_total_sess + s_extra_count
                total_student_present = present_count + s_extra_count
                total_student_absent = max(0, total_student_sessions - total_student_present)
                
                final_pct = round((total_student_present / total_student_sessions) * 100.0, 2) if total_student_sessions > 0 else 0.0
                is_defaulter = final_pct < settings.DEFAULTER_THRESHOLD_PERCENT if total_student_sessions > 0 else False

                if is_defaulter:
                    b_defaulters += 1
                b_pct_sum += final_pct

                photo_url = getattr(s, "photo_url", None)
                if not photo_url and getattr(s, "photo_urls", None) and len(s.photo_urls) > 0:
                    photo_url = s.photo_urls[0]

                # If attendance_type filter is REGULAR and user doesn't want extra
                if att_type_filter == "EXTRA_LECTURE" and s_extra_count == 0 and not reg_records:
                    continue

                b_students_summary.append({
                    "student_id": s.id,
                    "roll_number": s.roll_number,
                    "full_name": s.full_name,
                    "department": s.department,
                    "program": prog,
                    "semester": sem,
                    "division": sec,
                    "photo_url": photo_url,
                    
                    # Normal Sessions Breakdown
                    "normal_sessions": cur_total_sess,
                    "normal_present": present_count,
                    "normal_absent": absent_count,
                    "normal_percentage": normal_pct,
                    
                    # Extra Lectures
                    "extra_lectures": s_extra_count,
                    "extra_lecture_count": s_extra_count,
                    
                    # Combined Total Sessions & Final %
                    "total_sessions": total_student_sessions,
                    "total_present": total_student_present,
                    "total_absent": total_student_absent,
                    "final_percentage": final_pct,
                    "attendance_percentage": final_pct,
                    
                    # Status
                    "is_defaulter": is_defaulter,
                    "eligibility_status": "DEFAULTER" if is_defaulter else "ELIGIBLE",
                    
                    # Daily matrix & legacy compatibility
                    "present_count": total_student_present,
                    "absent_count": total_student_absent,
                    "total_recorded_attended": total_student_present,
                    "daily_status": daily_map
                })

            b_avg_pct = round(b_pct_sum / len(b_students), 2) if b_students else 0.0
            total_students_count += len(b_students)
            overall_defaulters += b_defaulters
            total_extra_lectures_conducted += b_extra_total

            batch_obj = {
                "batch_id": b_key,
                "batch_label": f"{prog} — {sem} — Division {sec}",
                "program": prog,
                "semester": sem,
                "division": sec,
                "division_code": sec,
                "division_name": f"Division {sec}",
                "total_enrolled": len(b_students),
                "total_sessions_conducted": b_sessions_count,
                "average_attendance_percentage": b_avg_pct,
                "defaulters_count": b_defaulters,
                "extra_lectures_count": b_extra_total,
                "students": b_students_summary
            }
            batches_list.append(batch_obj)

            div_key = f"{prog} {sem} Div {sec}" if len(batch_map) > 1 else sec
            divisions_data[div_key] = batch_obj

            if prog not in hierarchy:
                hierarchy[prog] = {}
            if sem not in hierarchy[prog]:
                hierarchy[prog][sem] = {}
            hierarchy[prog][sem][sec] = batch_obj

        all_students_flat = [s for b in batches_list for s in b["students"]]
        overall_avg_pct = round(sum(s["attendance_percentage"] for s in all_students_flat) / len(all_students_flat), 2) if all_students_flat else 0.0

        scope_parts = []
        if dept_val and dept_val.upper() != "ALL":
            scope_parts.append(dept_val)
        if programs_list:
            scope_parts.append("/".join(programs_list))
        if semesters_list:
            scope_parts.append("/".join(semesters_list))
        if divisions_list:
            scope_parts.append("Div " + ",".join(divisions_list))
        if c_ids_list and len(c_ids_list) == 1 and target_courses:
            scope_parts.append(f"{target_courses[0].code} - {target_courses[0].name}")

        scope_title = " — ".join(scope_parts) if scope_parts else "Institutional Attendance Master Dossier"

        return {
            "scope_title": scope_title,
            "class_id": c_ids_list[0] if len(c_ids_list) == 1 else None,
            "class_code": target_courses[0].code if (c_ids_list and len(c_ids_list) == 1 and target_courses) else "ALL",
            "class_name": target_courses[0].name if (c_ids_list and len(c_ids_list) == 1 and target_courses) else "All Academic Courses",
            "department": dept_val or "All Departments",
            "programs": list(hierarchy.keys()) if hierarchy else (programs_list or ["All Programs"]),
            "semesters": semesters_list or ["All Semesters"],
            "selected_divisions": divisions_list or list({b["division"] for b in batches_list}),
            "semester": "/".join(semesters_list) if semesters_list else "All Semesters",
            "selected_division": ",".join(divisions_list) if divisions_list else "ALL",
            "start_date": start_date.isoformat() if start_date else (unique_date_strs[0] if unique_date_strs else None),
            "end_date": end_date.isoformat() if end_date else date.today().isoformat(),
            "unique_dates": unique_date_strs,
            "total_sessions_conducted": total_sessions_count,
            "total_enrolled": total_students_count,
            "average_attendance_percentage": overall_avg_pct,
            "defaulters_count": overall_defaulters,
            "total_extra_lectures_conducted": total_extra_lectures_conducted,
            "hierarchy": hierarchy,
            "batches": batches_list,
            "divisions": divisions_data
        }

    # =========================================================================
    # INDIVIDUAL STUDENT DEEP-DIVE & BUNK LOG REPORTING
    # =========================================================================
    @staticmethod
    def get_student_detailed_report(
        db: Session,
        student_id: int,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> Dict[str, Any]:
        """
        Gathers complete attendance audit transcript for an individual student:
        - Profile Details
        - Normal Course Attendance KPIs (Conducted, Attended, Missed, %)
        - Extra Lecture Attendance (Outside course attendances separately tracked)
        - Subject-by-Subject Breakdown
        - Detailed Chronological Lecture-by-Lecture Log with Normal vs Extra Lecture tags
        """
        student = db.query(Student).filter(Student.id == student_id).first()
        if not student:
            return None

        # 1. Determine Applicable Normal Courses for this Student (Academic Curriculum Scope)
        s_prog = (getattr(student, "program", None) or "").strip().upper()
        s_sem = (student.semester or "").strip().upper()
        s_sec = (student.section or "").strip().upper()

        cq = db.query(ClassCourse)
        if student.department and student.department.upper() != "ALL":
            cq = cq.filter(ClassCourse.department == student.department)
        if student.program:
            cq = cq.filter(ClassCourse.program == student.program)
        if student.semester:
            cq = cq.filter(ClassCourse.semester == student.semester)
        if student.section:
            cq = cq.filter(ClassCourse.section.in_([student.section, "ALL", "BOTH", "*"]))
        batch_courses = cq.all()

        s_enrolled = [
            c for c in student.enrolled_classes
            if not s_prog 
            or not getattr(c, "program", None) 
            or c.program.strip().upper() in ["ALL", "*", "ANY", s_prog]
        ] if getattr(student, "enrolled_classes", None) and len(student.enrolled_classes) > 0 else []

        enrolled_courses = s_enrolled if s_enrolled else batch_courses
        enrolled_course_ids = [c.id for c in enrolled_courses]

        # 2. Fetch Normal Sessions for Enrolled Courses
        normal_session_query = db.query(AttendanceSession)
        if enrolled_course_ids:
            normal_session_query = normal_session_query.filter(AttendanceSession.class_id.in_(enrolled_course_ids))
        else:
            normal_session_query = normal_session_query.filter(AttendanceSession.id == -1)

        if start_date:
            normal_session_query = normal_session_query.filter(AttendanceSession.session_date >= start_date)
        if end_date:
            normal_session_query = normal_session_query.filter(AttendanceSession.session_date <= end_date)

        normal_sessions = normal_session_query.order_by(AttendanceSession.session_date.desc(), AttendanceSession.id.desc()).all()
        normal_session_ids = [s.id for s in normal_sessions]

        # 3. Fetch ALL attendance records for this student across ALL sessions in DB
        all_records_query = db.query(AttendanceRecord).join(AttendanceSession, AttendanceRecord.session_id == AttendanceSession.id).filter(
            AttendanceRecord.student_id == student.id
        )
        if start_date:
            all_records_query = all_records_query.filter(AttendanceSession.session_date >= start_date)
        if end_date:
            all_records_query = all_records_query.filter(AttendanceSession.session_date <= end_date)

        all_student_records = all_records_query.all()
        record_by_session_id: Dict[int, AttendanceRecord] = {r.session_id: r for r in all_student_records}

        # Build Map of Course ID to Course Object
        course_map: Dict[int, ClassCourse] = {c.id: c for c in enrolled_courses}
        # Also include any course referenced in all sessions the student has records for
        for r in all_student_records:
            if r.session and r.session.class_id and r.session.class_id not in course_map:
                c = db.query(ClassCourse).filter(ClassCourse.id == r.session.class_id).first()
                if c:
                    course_map[r.session.class_id] = c

        # 4. Subject-wise Tracking for Enrolled Courses
        subject_stats: Dict[int, Dict[str, Any]] = {}
        for c_id, c in course_map.items():
            teacher_name = c.teacher.full_name if (getattr(c, "teacher", None) and c.teacher) else "Faculty Coordinator"
            subject_stats[c_id] = {
                "course_id": c.id,
                "course_code": c.code,
                "course_name": c.name,
                "teacher_name": teacher_name,
                "department": c.department,
                "program": getattr(c, "program", "B.Tech"),
                "semester": c.semester,
                "section": c.section,
                "is_enrolled": c_id in enrolled_course_ids,
                "total_lectures": 0,
                "present_count": 0,
                "absent_count": 0,
                "extra_lecture_count": 0,
                "attendance_percentage": 0.0,
                "is_defaulter": False
            }

        # 5. Process Normal Course Sessions
        valid_normal_sessions = list(normal_sessions)

        lecture_history = []
        normal_total = len(valid_normal_sessions)
        normal_present = 0
        normal_bunk = 0
        normal_frozen = 0

        processed_session_ids = set()

        for sess in valid_normal_sessions:
            processed_session_ids.add(sess.id)
            c = course_map.get(sess.class_id)
            c_code = c.code if c else "GEN-101"
            c_name = c.name if c else (sess.session_name or "Academic Lecture")
            teacher_name = c.teacher.full_name if (c and getattr(c, "teacher", None) and c.teacher) else (sess.teacher.full_name if sess.teacher else "Faculty Coordinator")
            
            rec = record_by_session_id.get(sess.id)
            status = "ABSENT"
            confidence = 0.0
            verified_by = "Automated AI"
            is_extra = False

            if rec:
                status = rec.status
                confidence = rec.confidence_score or 0.0
                verified_by = getattr(rec, "verification_type", None) or "AI Ingestion"
                is_extra = bool(rec.is_extra_lecture or rec.attendance_type == "EXTRA_LECTURE" or rec.verification_type == "EXTRA_LECTURE")

            is_present = status in ["PRESENT", "LATE"]
            is_record_frozen = (status == "FROZEN")

            if is_record_frozen:
                normal_frozen += 1
            elif is_extra:
                # Extra lectures are tracked under extra_lectures_list, not normal present
                pass
            elif is_present:
                normal_present += 1
            else:
                normal_bunk += 1

            if sess.class_id in subject_stats:
                if not is_record_frozen:
                    subject_stats[sess.class_id]["total_lectures"] += 1
                    if is_present and not is_extra:
                        subject_stats[sess.class_id]["present_count"] += 1
                    elif not is_extra:
                        subject_stats[sess.class_id]["absent_count"] += 1

            s_date_str = sess.session_date.isoformat() if sess.session_date else date.today().isoformat()
            actual_time_str = format_ist_time(sess.created_at) if sess.created_at else (sess.start_time or "09:00 AM")
            scheduled_time_str = sess.start_time or (getattr(c, "start_time", None) if c else None)

            status_display = "FROZEN" if is_record_frozen else ("PRESENT" if is_present else "ABSENT")

            lecture_history.append({
                "session_id": sess.id,
                "date": s_date_str,
                "time": actual_time_str,
                "actual_time": actual_time_str,
                "scheduled_time": scheduled_time_str,
                "created_at": format_iso_utc(sess.created_at),
                "course_code": c_code,
                "course_name": c_name,
                "topic": sess.session_name or "Regular Classroom Lecture",
                "teacher_name": teacher_name,
                "status": status_display,
                "attendance_type": "EXTRA_LECTURE" if is_extra else "REGULAR",
                "is_extra_lecture": is_extra,
                "is_frozen": is_record_frozen,
                "is_bunk": not is_present and not is_record_frozen,
                "confidence_score": confidence,
                "verification_method": verified_by
            })

        # 6. Process Outside / Extra Lecture Records strictly from approved EXTRA_LECTURE records or outside normal sessions
        extra_records_query = db.query(AttendanceRecord).join(AttendanceSession, AttendanceRecord.session_id == AttendanceSession.id).filter(
            AttendanceRecord.student_id == student.id,
            AttendanceRecord.status.in_(["PRESENT", "LATE"]),
            (AttendanceRecord.is_extra_lecture == True) | 
            (AttendanceRecord.verification_type == "EXTRA_LECTURE") | 
            (AttendanceRecord.attendance_type == "EXTRA_LECTURE") |
            (~AttendanceRecord.session_id.in_(normal_session_ids) if normal_session_ids else True)
        )
        if start_date:
            extra_records_query = extra_records_query.filter(AttendanceSession.session_date >= start_date)
        if end_date:
            extra_records_query = extra_records_query.filter(AttendanceSession.session_date <= end_date)

        extra_records = extra_records_query.all()
        
        extra_lectures_list = []
        seen_extra_session_ids = set()
        for rec in extra_records:
            sess = rec.session
            if not sess or sess.id in seen_extra_session_ids:
                continue
            seen_extra_session_ids.add(sess.id)

            c = course_map.get(sess.class_id)
            c_code = c.code if c else "EXT-101"
            c_name = c.name if c else (sess.session_name or "Extra Lecture")
            teacher_name = c.teacher.full_name if (c and getattr(c, "teacher", None) and c.teacher) else (sess.teacher.full_name if sess.teacher else "Faculty Coordinator")
            
            s_date_str = sess.session_date.isoformat() if sess.session_date else date.today().isoformat()
            actual_time_str = format_ist_time(sess.created_at) if sess.created_at else (sess.start_time or "09:00 AM")
            scheduled_time_str = sess.start_time or (getattr(c, "start_time", None) if c else None)

            extra_item = {
                "session_id": sess.id,
                "date": s_date_str,
                "time": actual_time_str,
                "actual_time": actual_time_str,
                "scheduled_time": scheduled_time_str,
                "created_at": format_iso_utc(sess.created_at),
                "course_code": c_code,
                "course_name": c_name,
                "topic": sess.session_name or "Extra Lecture",
                "teacher_name": teacher_name,
                "status": "PRESENT",
                "attendance_type": "EXTRA_LECTURE",
                "is_extra_lecture": True,
                "is_bunk": False,
                "confidence_score": rec.confidence_score or 95.0,
                "verification_method": "Extra Lecture Approved"
            }
            extra_lectures_list.append(extra_item)
            
            if sess.id not in processed_session_ids:
                lecture_history.append(extra_item)
                processed_session_ids.add(sess.id)

            if sess.class_id in subject_stats:
                subject_stats[sess.class_id]["extra_lecture_count"] += 1
            else:
                subject_stats[sess.class_id] = {
                    "course_id": c.id if c else sess.class_id,
                    "course_code": c_code,
                    "course_name": c_name,
                    "teacher_name": teacher_name,
                    "department": c.department if c else "Academic",
                    "program": getattr(c, "program", "B.Tech") if c else "Academic",
                    "semester": c.semester if c else "N/A",
                    "section": c.section if c else "N/A",
                    "is_enrolled": sess.class_id in enrolled_course_ids,
                    "total_lectures": 0,
                    "present_count": 0,
                    "absent_count": 0,
                    "extra_lecture_count": 1,
                    "attendance_percentage": 100.0,
                    "is_defaulter": False
                }

        # Sort lecture history chronologically descending
        lecture_history.sort(key=lambda x: (x["date"], x["time"]), reverse=True)

        # Subject-wise calculations
        subjects_list = []
        for c_id, s_info in subject_stats.items():
            tot = s_info["total_lectures"]
            pr = s_info["present_count"]
            pct = round((pr / tot) * 100.0, 2) if tot > 0 else 100.0
            s_info["attendance_percentage"] = pct
            s_info["is_defaulter"] = pct < settings.DEFAULTER_THRESHOLD_PERCENT if tot > 0 else False
            if s_info["is_enrolled"] or s_info["total_lectures"] > 0 or s_info["extra_lecture_count"] > 0:
                subjects_list.append(s_info)

        normal_eligible = max(0, normal_total - normal_frozen)
        normal_pct = round((normal_present / normal_eligible) * 100.0, 2) if normal_eligible > 0 else 0.0
        extra_count = len(extra_lectures_list)

        # EXACT REQUIRED COMBINED LOGIC (Exempting Frozen Sessions):
        # Total Sessions = Normal Eligible Sessions + Extra Lectures
        # Total Present = Normal Present + Extra Lectures
        # Total Absent = Total Sessions - Total Present
        # Final Attendance % = (Total Present / Total Sessions) * 100
        total_sessions = normal_eligible + extra_count
        total_present = normal_present + extra_count
        total_absent = max(0, total_sessions - total_present)

        final_pct = round((total_present / total_sessions) * 100.0, 2) if total_sessions > 0 else 0.0
        is_overall_defaulter = final_pct < settings.DEFAULTER_THRESHOLD_PERCENT if total_sessions > 0 else False

        photo_url = getattr(student, "photo_url", None)
        if not photo_url and getattr(student, "photo_urls", None) and len(student.photo_urls) > 0:
            photo_url = student.photo_urls[0]

        has_biometrics = bool(student.face_embedding or (getattr(student, "photo_urls", None) and len(student.photo_urls) > 0))
        photo_count = len(student.photo_urls) if getattr(student, "photo_urls", None) else (1 if photo_url else 0)

        is_student_frozen = bool(getattr(student, "is_frozen", False) or getattr(student, "attendance_status", "ACTIVE") == "FROZEN")

        return {
            "student_id": student.id,
            "roll_number": student.roll_number,
            "full_name": student.full_name,
            "email": student.email,
            "department": student.department,
            "program": getattr(student, "program", "B.Tech"),
            "semester": student.semester,
            "division": (student.section or "A").strip().upper(),
            "academic_year": getattr(student, "academic_year", "2026-27") or "2026-27",
            "admission_year": getattr(student, "admission_year", 2023) or 2023,
            "batch": getattr(student, "batch", "2023-2027") or "2023-2027",
            "mobile_number": getattr(student, "mobile_number", None),
            "status": getattr(student, "status", "Active") or "Active",
            "attendance_status": "FROZEN" if is_student_frozen else "ACTIVE",
            "is_frozen": is_student_frozen,
            "frozen_at": format_iso_utc(student.frozen_at) if getattr(student, "frozen_at", None) else None,
            "freeze_reason": getattr(student, "freeze_reason", None),
            "freeze_until": student.freeze_until.strftime("%Y-%m-%d") if getattr(student, "freeze_until", None) else None,
            "biometric_status": "Enrolled" if has_biometrics else "Pending",
            "biometric_enrolled": has_biometrics,
            "photo_count": photo_count,
            "photo_url": photo_url,
            "defaulter_threshold": settings.DEFAULTER_THRESHOLD_PERCENT,
            
            # Normal Sessions Breakdown
            "normal_sessions": normal_eligible,
            "normal_conducted": normal_total,
            "normal_eligible": normal_eligible,
            "normal_present": normal_present,
            "normal_absent": normal_bunk,
            "normal_frozen": normal_frozen,
            "normal_percentage": normal_pct,
            "total_frozen": normal_frozen,
            "total_lectures_conducted": normal_total,
            "total_lectures_attended": normal_present,
            "total_lectures_missed": normal_bunk,
            "normal_attendance_percentage": normal_pct,
            
            # Extra Lectures
            "extra_lectures": extra_lectures_list,
            "extra_lecture_count": extra_count,
            "extra_lectures_count": extra_count,
            
            # Combined Total Sessions
            "total_sessions": total_sessions,
            "total_present": total_present,
            "total_absent": total_absent,
            "final_percentage": final_pct,
            "final_attendance_percentage": final_pct,
            "overall_attendance_percentage": final_pct,
            "attendance_percentage": final_pct,
            "total_recorded_attendance": total_present,
            
            # Eligibility / Defaulter
            "is_defaulter": is_overall_defaulter,
            "eligibility_status": "DEFAULTER" if is_overall_defaulter else "ELIGIBLE",
            
            # Lists
            "subjects_breakdown": subjects_list,
            "lecture_history": lecture_history
        }

    @staticmethod
    def export_student_pdf(
        db: Session,
        student_id: int,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> str:
        """
        Exports official Student Attendance Transcript & Academic Audit Dossier PDF.
        Includes Compact Profile Photo, 3-Tier KPIs, Subject Breakdown, Extra Lectures,
        Chronological Timeline with Actual Timestamps, Repeating Table Headers,
        and NumberedCanvas ('Page X of Y').
        """
        data = ReportService.get_student_detailed_report(db, student_id, start_date, end_date)
        if not data:
            raise ValueError("Student not found")

        filename = f"Student_Attendance_{data['roll_number']}_{uuid.uuid4().hex[:6]}.pdf"
        filepath = str(settings.REPORTS_DIR / filename)

        doc = SimpleDocTemplate(
            filepath,
            pagesize=portrait(A4),
            rightMargin=28,
            leftMargin=28,
            topMargin=38,
            bottomMargin=42
        )

        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "DocTitle",
            parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=13,
            leading=16,
            textColor=colors.HexColor("#0F172A"),
            alignment=1,
            spaceAfter=2
        )

        subtitle_style = ParagraphStyle(
            "DocSub",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=11,
            textColor=colors.HexColor("#64748B"),
            alignment=1,
            spaceAfter=6
        )

        section_heading = ParagraphStyle(
            "SecHead",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=9,
            leading=12,
            textColor=colors.HexColor("#1E3A8A"),
            spaceBefore=6,
            spaceAfter=3
        )

        tbl_hdr = ParagraphStyle("THdr", fontName="Helvetica-Bold", fontSize=7, leading=8.5, textColor=colors.white, alignment=1)
        tbl_cell_c = ParagraphStyle("TCellC", fontName="Helvetica", fontSize=7, leading=8.5, textColor=colors.HexColor("#1E293B"), alignment=1)
        tbl_cell_l = ParagraphStyle("TCellL", fontName="Helvetica", fontSize=7, leading=8.5, textColor=colors.HexColor("#1E293B"), alignment=0)

        elements = []

        # Title Banner
        elements.append(Paragraph("VisionAttend Pro — Official Student Attendance Transcript", title_style))
        elements.append(Paragraph(f"Academic Biometric Audit Record &bull; Session Year {data.get('academic_year', '2026-27')}", subtitle_style))

        # Compact Photo Loader (~30mm)
        photo_element = None
        if data.get("photo_url"):
            p_url = data["photo_url"]
            local_path = None
            if p_url.startswith("/uploads/"):
                local_path = str(settings.UPLOAD_DIR / p_url.replace("/uploads/", ""))
            elif p_url.startswith("uploads/"):
                local_path = str(settings.PROJECT_ROOT / p_url)
            else:
                local_path = p_url

            if local_path and os.path.exists(local_path):
                try:
                    photo_element = RLImage(local_path, width=0.95*inch, height=0.95*inch, kind='proportional')
                except Exception:
                    photo_element = None

        # Student Profile Banner Table
        is_def = data.get("is_defaulter", False)
        status_color = "#B91C1C" if is_def else "#15803D"
        status_label = "DEFAULTER RISK (<75%)" if is_def else "ELIGIBLE (>=75%)"

        info_col1 = [
            Paragraph(f"<b>Student Name:</b> <font color='#0F172A'><b>{data['full_name']}</b></font>", tbl_cell_l),
            Paragraph(f"<b>Roll Number:</b> <font color='#4338CA'><b>{data['roll_number']}</b></font>", tbl_cell_l),
            Paragraph(f"<b>Program / Degree:</b> {data['program']}", tbl_cell_l),
            Paragraph(f"<b>Semester & Div:</b> {data['semester']} • Div {data['division']}", tbl_cell_l)
        ]
        info_col2 = [
            Paragraph(f"<b>Department:</b> {data['department']}", tbl_cell_l),
            Paragraph(f"<b>Academic Year:</b> {data.get('academic_year', '2026-27')}", tbl_cell_l),
            Paragraph(f"<b>Biometrics:</b> <font color='{'#15803D' if data.get('biometric_enrolled') else '#D97706'}'><b>{data.get('biometric_status', 'Enrolled')}</b></font>", tbl_cell_l),
            Paragraph(f"<b>Audit Status:</b> <font color='{status_color}'><b>{status_label}</b></font>", tbl_cell_l)
        ]

        if photo_element:
            prof_rows = [[
                photo_element,
                info_col1,
                info_col2
            ]]
            prof_tbl = Table(prof_rows, colWidths=[1.1*inch, 3.2*inch, 3.3*inch])
        else:
            prof_rows = [[
                info_col1,
                info_col2
            ]]
            prof_tbl = Table(prof_rows, colWidths=[3.8*inch, 3.8*inch])

        prof_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
            ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#CBD5E1")),
            ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (0, 0), "CENTER"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(prof_tbl)
        elements.append(Spacer(1, 6))

        # 3-Tier KPI Summary Table
        norm_color = "#15803D" if data.get('normal_percentage', 0) >= 75 else "#B91C1C"
        kpi_rows = [
            [
                Paragraph("<b>REGULAR SESSIONS</b>", tbl_hdr),
                Paragraph("<b>REGULAR ATTENDED</b>", tbl_hdr),
                Paragraph("<b>REGULAR MISSED</b>", tbl_hdr),
                Paragraph("<b>NORMAL RATE %</b>", tbl_hdr),
                Paragraph("<b>EXTRA LECTURES</b>", tbl_hdr),
                Paragraph("<b>TOTAL SESSIONS</b>", tbl_hdr),
                Paragraph("<b>TOTAL PRESENT</b>", tbl_hdr),
                Paragraph("<b>FINAL ATT. %</b>", tbl_hdr),
                Paragraph("<b>FINAL STATUS</b>", tbl_hdr)
            ],
            [
                Paragraph(f"<font size=9.5><b>{data['normal_sessions']}</b></font>", tbl_cell_c),
                Paragraph(f"<font size=9.5 color='#15803D'><b>{data['normal_present']}</b></font>", tbl_cell_c),
                Paragraph(f"<font size=9.5 color='#B91C1C'><b>{data['normal_absent']}</b></font>", tbl_cell_c),
                Paragraph(f"<font size=9.5 color='{norm_color}'><b>{data['normal_percentage']}%</b></font>", tbl_cell_c),
                Paragraph(f"<font size=9.5 color='#D97706'><b>{data['extra_lecture_count']}</b></font>", tbl_cell_c),
                Paragraph(f"<font size=9.5 color='#1E3A8A'><b>{data['total_sessions']}</b></font>", tbl_cell_c),
                Paragraph(f"<font size=9.5 color='#15803D'><b>{data['total_present']}</b></font>", tbl_cell_c),
                Paragraph(f"<font size=9.5 color='{status_color}'><b>{data['final_percentage']}%</b></font>", tbl_cell_c),
                Paragraph(f"<font size=8.5 color='{status_color}'><b>{data['eligibility_status']}</b></font>", tbl_cell_c)
            ]
        ]
        kpi_tbl = Table(kpi_rows, colWidths=[0.84*inch]*7 + [0.86*inch, 0.86*inch])
        kpi_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
            ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#F1F5F9")),
            ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#94A3B8")),
            ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elements.append(kpi_tbl)
        elements.append(Spacer(1, 6))

        # Section 1: Subject-Wise Attendance Breakdown Table
        elements.append(Paragraph("<b>1. Subject-Wise Attendance Breakdown</b>", section_heading))
        subj_rows = [
            [
                Paragraph("<b>Code</b>", tbl_hdr),
                Paragraph("<b>Subject / Course Name</b>", tbl_hdr),
                Paragraph("<b>Faculty In-Charge</b>", tbl_hdr),
                Paragraph("<b>Conducted</b>", tbl_hdr),
                Paragraph("<b>Attended</b>", tbl_hdr),
                Paragraph("<b>Missed</b>", tbl_hdr),
                Paragraph("<b>Extra Credit</b>", tbl_hdr),
                Paragraph("<b>Att. %</b>", tbl_hdr),
                Paragraph("<b>Status</b>", tbl_hdr),
            ]
        ]

        t1_style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 2.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
            ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#94A3B8")),
        ]

        for idx, sub in enumerate(data.get("subjects_breakdown", []), 1):
            s_is_def = sub.get("is_defaulter", False)
            s_col = "#991B1B" if s_is_def else "#15803D"
            extra_credit = f"+{sub['extra_lecture_count']}" if sub.get('extra_lecture_count', 0) > 0 else "-"
            subj_rows.append([
                Paragraph(sub["course_code"], tbl_cell_c),
                Paragraph(sub["course_name"], tbl_cell_l),
                Paragraph(sub["teacher_name"], tbl_cell_l),
                Paragraph(str(sub["total_lectures"]), tbl_cell_c),
                Paragraph(str(sub["present_count"]), tbl_cell_c),
                Paragraph(str(sub["absent_count"]), tbl_cell_c),
                Paragraph(extra_credit, tbl_cell_c),
                Paragraph(f"<font color='{s_col}'><b>{sub['attendance_percentage']}%</b></font>", tbl_cell_c),
                Paragraph(f"<font color='{s_col}'><b>{'DEFAULTER' if s_is_def else 'ELIGIBLE'}</b></font>", tbl_cell_c)
            ])
            if s_is_def:
                t1_style.append(("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#FEE2E2")))
            elif idx % 2 == 0:
                t1_style.append(("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#F8FAFC")))

        subj_tbl = Table(
            subj_rows,
            colWidths=[0.75*inch, 2.3*inch, 1.45*inch, 0.55*inch, 0.55*inch, 0.5*inch, 0.55*inch, 0.55*inch, 0.65*inch]
        )
        subj_tbl.setStyle(TableStyle(t1_style))
        elements.append(subj_tbl)
        elements.append(Spacer(1, 6))

        # Section 2: Approved Extra Lecture Records (if any)
        if data.get("extra_lectures") and len(data["extra_lectures"]) > 0:
            elements.append(Paragraph(f"<b>2. Approved Extra Lecture Records ({len(data['extra_lectures'])})</b>", section_heading))
            extra_rows = [
                [
                    Paragraph("<b>Date & Time</b>", tbl_hdr),
                    Paragraph("<b>Course Code & Name</b>", tbl_hdr),
                    Paragraph("<b>Lecture Topic</b>", tbl_hdr),
                    Paragraph("<b>Faculty</b>", tbl_hdr),
                    Paragraph("<b>Status</b>", tbl_hdr),
                    Paragraph("<b>Credit & Approval</b>", tbl_hdr)
                ]
            ]
            t_extra_style = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#B45309")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 2.5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#FDE68A")),
                ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#D97706")),
            ]
            for e_idx, ex in enumerate(data["extra_lectures"], 1):
                extra_rows.append([
                    Paragraph(f"{ex.get('date', '')} {ex.get('actual_time') or ex.get('time', '')}", tbl_cell_c),
                    Paragraph(f"<b>{ex.get('course_code', '')}</b>: {ex.get('course_name', '')}", tbl_cell_l),
                    Paragraph(ex.get("topic", "Extra Lecture"), tbl_cell_l),
                    Paragraph(ex.get("teacher_name", "Faculty Coordinator"), tbl_cell_l),
                    Paragraph("<font color='#15803D'><b>PRESENT</b></font>", tbl_cell_c),
                    Paragraph("<font color='#B45309'><b>+1 Extra Credit (Approved)</b></font>", tbl_cell_c)
                ])
                t_extra_style.append(("BACKGROUND", (0, e_idx), (-1, e_idx), colors.HexColor("#FEF3C7")))

            extra_tbl = Table(extra_rows, colWidths=[1.1*inch, 1.9*inch, 1.8*inch, 1.3*inch, 0.7*inch, 1.05*inch])
            extra_tbl.setStyle(TableStyle(t_extra_style))
            elements.append(extra_tbl)
            elements.append(Spacer(1, 6))

        # Section 3: Chronological Timeline (with repeatRows=1)
        elements.append(Paragraph("<b>3. Chronological Lecture-by-Lecture Timeline & Audit Log</b>", section_heading))
        log_rows = [
            [
                Paragraph("<b>Date & Time</b>", tbl_hdr),
                Paragraph("<b>Subject / Course</b>", tbl_hdr),
                Paragraph("<b>Lecture Topic</b>", tbl_hdr),
                Paragraph("<b>Faculty</b>", tbl_hdr),
                Paragraph("<b>Type</b>", tbl_hdr),
                Paragraph("<b>Status</b>", tbl_hdr),
                Paragraph("<b>Source / Verification</b>", tbl_hdr)
            ]
        ]

        t2_style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#94A3B8")),
        ]

        for idx, log in enumerate(data.get("lecture_history", []), 1):
            is_bunk = log.get("is_bunk", False)
            is_extra = log.get("is_extra_lecture", False) or log.get("attendance_type") == "EXTRA_LECTURE"
            act_time = log.get("actual_time") or log.get("time", "")

            if is_extra:
                type_text = "<font color='#D97706'><b>EXTRA</b></font>"
                status_text = "<font color='#15803D'><b>PRESENT</b></font>"
            else:
                type_text = "<font color='#475569'>NORMAL</font>"
                status_text = "<font color='#B91C1C'><b>MISSED</b></font>" if is_bunk else "<font color='#15803D'><b>PRESENT</b></font>"

            log_rows.append([
                Paragraph(f"{log.get('date', '')} {act_time}", tbl_cell_c),
                Paragraph(f"<b>{log.get('course_code', '')}</b>", tbl_cell_l),
                Paragraph(log.get("topic", "Lecture"), tbl_cell_l),
                Paragraph(log.get("teacher_name", "Faculty"), tbl_cell_l),
                Paragraph(type_text, tbl_cell_c),
                Paragraph(status_text, tbl_cell_c),
                Paragraph(log.get("verification_method", "AI Ingestion"), tbl_cell_c)
            ])
            if is_extra:
                t2_style.append(("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#FEF3C7")))
            elif is_bunk:
                t2_style.append(("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#FEE2E2")))
            elif idx % 2 == 0:
                t2_style.append(("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#F8FAFC")))

        log_tbl = Table(
            log_rows,
            colWidths=[1.1*inch, 1.1*inch, 1.9*inch, 1.3*inch, 0.65*inch, 0.65*inch, 1.15*inch],
            repeatRows=1
        )
        log_tbl.setStyle(TableStyle(t2_style))
        elements.append(log_tbl)
        elements.append(Spacer(1, 10))

        # Official Institutional Sign-Off Table
        signoff_data = [
            [
                Paragraph("___________________________<br/><b>Faculty Advisor / Mentor</b>", tbl_cell_c),
                Paragraph("___________________________<br/><b>Head of Department (HOD)</b>", tbl_cell_c),
                Paragraph("___________________________<br/><b>Student Acknowledgment</b>", tbl_cell_c)
            ]
        ]
        signoff_tbl = Table(signoff_data, colWidths=[2.6*inch, 2.6*inch, 2.65*inch])
        signoff_tbl.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("TOPPADDING", (0, 0), (-1, -1), 10),
        ]))
        elements.append(KeepTogether(signoff_tbl))

        doc.build(elements, canvasmaker=NumberedCanvas)
        return filepath

    # =========================================================================
    # ADVANCED EXCEL EXPORT (MULTI-SHEET WITH EXTRA LECTURES)
    # =========================================================================
    @staticmethod
    def export_advanced_excel(
        db: Session,
        class_id: Optional[int] = None,
        class_ids: Optional[Any] = None,
        department: Optional[str] = None,
        departments: Optional[Any] = None,
        program: Optional[str] = None,
        programs: Optional[Any] = None,
        semester: Optional[str] = None,
        semesters: Optional[Any] = None,
        division: Optional[str] = None,
        divisions: Optional[Any] = None,
        attendance_type: Optional[str] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> str:
        data = ReportService.get_advanced_report_data(
            db=db,
            class_id=class_id,
            class_ids=class_ids,
            department=department or departments,
            programs=programs or program,
            semesters=semesters or semester,
            divisions=divisions or division,
            attendance_type=attendance_type,
            start_date=start_date,
            end_date=end_date
        )

        wb = openpyxl.Workbook()

        navy_header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        accent_blue_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        amber_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
        section_banner_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        amber_light_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        light_gray_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

        white_title_font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
        white_header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        white_bold_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        bold_font = Font(name="Segoe UI", size=10, bold=True)
        regular_font = Font(name="Segoe UI", size=9.5)
        defaulter_font = Font(name="Segoe UI", size=9.5, bold=True, color="991B1B")

        thin_border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1")
        )

        unique_dates = data.get("unique_dates", [])
        batches = data.get("batches", [])

        # Sheet 1: Executive Summary
        ws_sum = wb.active
        ws_sum.title = "Executive Summary"
        ws_sum.views.sheetView[0].showGridLines = True

        ws_sum.merge_cells("A1:I1")
        ws_sum["A1"] = f"VisionAttend Pro — Institutional Attendance Dossier"
        ws_sum["A1"].font = white_title_font
        ws_sum["A1"].fill = navy_header_fill
        ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_sum.row_dimensions[1].height = 38

        ws_sum.merge_cells("A2:I2")
        ws_sum["A2"] = f"Scope: {data['scope_title']} | Date Range: {data['start_date'] or 'Start'} to {data['end_date'] or 'Present'} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws_sum["A2"].font = Font(name="Segoe UI", size=9, italic=True, color="64748B")
        ws_sum["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws_sum.row_dimensions[2].height = 20

        meta_items = [
            ("Academic Scope:", data["scope_title"], "Department:", str(data["department"])),
            ("Programs / Degrees:", ", ".join(data.get("programs", [])), "Semesters / Batches:", str(data["semester"])),
            ("Total Sessions Held:", str(data["total_sessions_conducted"]), "Overall Average %:", f"{data['average_attendance_percentage']}%"),
            ("Total Enrolled Students:", str(data["total_enrolled"]), "Extra Lectures Conducted:", f"{data.get('total_extra_lectures_conducted', 0)} sessions"),
            ("Defaulters (<75%):", f"{data['defaulters_count']} students", "Attendance Types Included:", "Normal & Extra Lecture")
        ]

        curr_row = 4
        for c1_label, c1_val, c2_label, c2_val in meta_items:
            ws_sum.cell(row=curr_row, column=1, value=c1_label).font = bold_font
            ws_sum.cell(row=curr_row, column=2, value=c1_val).font = regular_font
            ws_sum.cell(row=curr_row, column=5, value=c2_label).font = bold_font
            ws_sum.cell(row=curr_row, column=6, value=c2_val).font = regular_font
            ws_sum.row_dimensions[curr_row].height = 19
            curr_row += 1

        curr_row += 1

        ws_sum.merge_cells(f"A{curr_row}:I{curr_row}")
        ws_sum[f"A{curr_row}"] = "PROGRAM & DIVISION-WISE ATTENDANCE BREAKDOWN"
        ws_sum[f"A{curr_row}"].font = white_bold_font
        ws_sum[f"A{curr_row}"].fill = section_banner_fill
        ws_sum[f"A{curr_row}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws_sum.row_dimensions[curr_row].height = 24
        curr_row += 1

        div_headers = [
            "Program", "Semester", "Division", "Enrolled", "Normal Sessions",
            "Normal Avg %", "Extra Lectures", "Defaulters (<75%)", "Batch Status"
        ]
        for c_idx, h in enumerate(div_headers, 1):
            cell = ws_sum.cell(row=curr_row, column=c_idx, value=h)
            cell.font = white_header_font
            cell.fill = accent_blue_fill
            cell.alignment = Alignment(horizontal="center" if c_idx > 3 else "left", vertical="center")
            cell.border = thin_border
        ws_sum.row_dimensions[curr_row].height = 24
        curr_row += 1

        for b in batches:
            ws_sum.cell(row=curr_row, column=1, value=b["program"]).alignment = Alignment(horizontal="left", vertical="center")
            ws_sum.cell(row=curr_row, column=2, value=b["semester"]).alignment = Alignment(horizontal="left", vertical="center")
            ws_sum.cell(row=curr_row, column=3, value=f"Division {b['division']}").alignment = Alignment(horizontal="left", vertical="center")
            ws_sum.cell(row=curr_row, column=4, value=b["total_enrolled"]).alignment = Alignment(horizontal="center", vertical="center")
            ws_sum.cell(row=curr_row, column=5, value=b["total_sessions_conducted"]).alignment = Alignment(horizontal="center", vertical="center")
            
            avg_cell = ws_sum.cell(row=curr_row, column=6, value=f"{b['average_attendance_percentage']}%")
            avg_cell.alignment = Alignment(horizontal="center", vertical="center")
            avg_cell.font = Font(name="Segoe UI", size=9.5, bold=True, color="15803D" if b["average_attendance_percentage"] >= 75 else "B91C1C")

            extra_cell = ws_sum.cell(row=curr_row, column=7, value=b.get("extra_lectures_count", 0))
            extra_cell.alignment = Alignment(horizontal="center", vertical="center")
            extra_cell.font = Font(name="Segoe UI", size=9.5, bold=True, color="D97706")

            def_cell = ws_sum.cell(row=curr_row, column=8, value=b["defaulters_count"])
            def_cell.alignment = Alignment(horizontal="center", vertical="center")
            def_cell.font = Font(name="Segoe UI", size=9.5, bold=True, color="B91C1C" if b["defaulters_count"] > 0 else "15803D")

            status_cell = ws_sum.cell(row=curr_row, column=9, value="Healthy" if b["average_attendance_percentage"] >= 75 else "Action Required")
            status_cell.alignment = Alignment(horizontal="center", vertical="center")
            status_cell.fill = green_fill if b["average_attendance_percentage"] >= 75 else red_fill

            for col in range(1, 10):
                ws_sum.cell(row=curr_row, column=col).border = thin_border
            ws_sum.row_dimensions[curr_row].height = 20
            curr_row += 1

        for col in range(1, 10):
            ws_sum.column_dimensions[get_column_letter(col)].width = 18

        # Sheets 2+: Batch-wise Matrices
        used_sheet_titles = set(["Executive Summary"])

        for b in batches:
            prog_abbr = b['program'].replace('.','').replace(' ','')[:6]
            sem_abbr = b['semester'].replace('Semester ', 'S').replace(' ', '')
            raw_title = f"{prog_abbr} {sem_abbr} Div {b['division']}"
            sheet_title = raw_title[:31]
            if sheet_title in used_sheet_titles:
                sheet_title = f"{sheet_title[:28]}_{len(used_sheet_titles)}"
            used_sheet_titles.add(sheet_title)

            ws = wb.create_sheet(title=sheet_title)
            ws.views.sheetView[0].showGridLines = True

            # Banner
            num_date_cols = len(unique_dates)
            max_col = 8 + num_date_cols
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
            ws["A1"] = f"{b['batch_label']} — Attendance Master Matrix"
            ws["A1"].font = white_bold_font
            ws["A1"].fill = navy_header_fill
            ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[1].height = 28

            headers = [
                "Roll Number", "Student Full Name", "Program", "Sem", "Div",
                "Normal Sessions", "Normal Present", "Normal Absent", "Normal %",
                "Extra Lectures", "Total Sessions", "Total Present", "Total Absent",
                "Final Attendance %", "Status"
            ]
            for d in unique_dates:
                headers.append(d)

            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col_idx, value=h)
                cell.font = white_header_font
                cell.fill = accent_blue_fill
                cell.alignment = Alignment(horizontal="center" if col_idx > 2 else "left", vertical="center")
                cell.border = thin_border
            ws.row_dimensions[2].height = 24

            for r_idx, s in enumerate(b["students"], 3):
                is_def = s["is_defaulter"]
                ws.cell(row=r_idx, column=1, value=s["roll_number"]).alignment = Alignment(horizontal="left", vertical="center")
                ws.cell(row=r_idx, column=2, value=s["full_name"]).alignment = Alignment(horizontal="left", vertical="center")
                ws.cell(row=r_idx, column=3, value=s["program"]).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=r_idx, column=4, value=s["semester"]).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=r_idx, column=5, value=s["division"]).alignment = Alignment(horizontal="center", vertical="center")
                
                # Normal Breakdown
                ws.cell(row=r_idx, column=6, value=s["normal_sessions"]).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=r_idx, column=7, value=s["normal_present"]).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=r_idx, column=8, value=s["normal_absent"]).alignment = Alignment(horizontal="center", vertical="center")
                
                norm_pct_cell = ws.cell(row=r_idx, column=9, value=f"{s['normal_percentage']}%")
                norm_pct_cell.alignment = Alignment(horizontal="center", vertical="center")
                norm_pct_cell.font = Font(name="Segoe UI", size=9, bold=True, color="15803D" if s["normal_percentage"] >= 75 else "B91C1C")

                # Extra Lectures
                extra_cell = ws.cell(row=r_idx, column=10, value=s.get("extra_lectures", s.get("extra_lecture_count", 0)))
                extra_cell.alignment = Alignment(horizontal="center", vertical="center")
                extra_cell.font = Font(name="Segoe UI", size=9, bold=True, color="D97706")

                # Combined Total
                ws.cell(row=r_idx, column=11, value=s["total_sessions"]).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=r_idx, column=12, value=s["total_present"]).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=r_idx, column=13, value=s["total_absent"]).alignment = Alignment(horizontal="center", vertical="center")

                final_pct_cell = ws.cell(row=r_idx, column=14, value=f"{s['final_percentage']}%")
                final_pct_cell.alignment = Alignment(horizontal="center", vertical="center")
                final_pct_cell.font = Font(name="Segoe UI", size=9.5, bold=True, color="991B1B" if is_def else "15803D")

                st_cell = ws.cell(row=r_idx, column=15, value="DEFAULTER" if is_def else "ELIGIBLE")
                st_cell.alignment = Alignment(horizontal="center", vertical="center")
                st_cell.font = Font(name="Segoe UI", size=8.5, bold=True, color="991B1B" if is_def else "15803D")
                st_cell.fill = red_fill if is_def else green_fill

                # Date marks (attended/total per day)
                for d_idx, d_str in enumerate(unique_dates, 16):
                    day_data = s.get("daily_status", {}).get(d_str)
                    if day_data and isinstance(day_data, dict):
                        mark = f"{day_data.get('attended', 0)}/{day_data.get('total', 0)}"
                        is_present = day_data.get("attended", 0) > 0
                    else:
                        # Legacy fallback for P/A strings
                        mark = day_data if day_data else "A"
                        is_present = mark == "P"
                    d_cell = ws.cell(row=r_idx, column=d_idx, value=mark)
                    d_cell.alignment = Alignment(horizontal="center", vertical="center")
                    d_cell.font = Font(name="Segoe UI", size=8.5, bold=True, color="15803D" if is_present else "B91C1C")
                    d_cell.fill = green_fill if is_present else light_gray_fill

                for col in range(1, len(headers) + 1):
                    ws.cell(row=r_idx, column=col).border = thin_border
                ws.row_dimensions[r_idx].height = 19

            ws.column_dimensions["A"].width = 15
            ws.column_dimensions["B"].width = 24
            for c in ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"]:
                ws.column_dimensions[c].width = 13
            for c_idx in range(16, len(headers) + 1):
                ws.column_dimensions[get_column_letter(c_idx)].width = 12

        # Dedicated Sheet for Extra Lecture Records (if any)
        extra_records_query = db.query(AttendanceRecord).join(AttendanceSession, AttendanceRecord.session_id == AttendanceSession.id).join(Student, AttendanceRecord.student_id == Student.id).filter(
            (AttendanceRecord.is_extra_lecture == True) | 
            (AttendanceRecord.verification_type == "EXTRA_LECTURE") | 
            (AttendanceRecord.attendance_type == "EXTRA_LECTURE")
        )
        if start_date:
            extra_records_query = extra_records_query.filter(AttendanceSession.session_date >= start_date)
        if end_date:
            extra_records_query = extra_records_query.filter(AttendanceSession.session_date <= end_date)

        all_extra_records = extra_records_query.order_by(AttendanceSession.session_date.desc(), AttendanceRecord.id.desc()).all()

        if all_extra_records:
            ws_extra = wb.create_sheet(title="Extra Lecture Records")
            ws_extra.views.sheetView[0].showGridLines = True

            ws_extra.merge_cells("A1:J1")
            ws_extra["A1"] = "EXTRA LECTURE ATTENDANCE AUDIT LOG"
            ws_extra["A1"].font = white_bold_font
            ws_extra["A1"].fill = amber_fill
            ws_extra["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws_extra.row_dimensions[1].height = 28

            extra_hdrs = [
                "Date", "Student Name", "Roll Number", "Student Academic Context",
                "Attended Course Code", "Attended Course Name", "Lecture Topic",
                "Faculty", "Status", "Attendance Type"
            ]
            for c_idx, h in enumerate(extra_hdrs, 1):
                c_cell = ws_extra.cell(row=2, column=c_idx, value=h)
                c_cell.font = white_header_font
                c_cell.fill = navy_header_fill
                c_cell.alignment = Alignment(horizontal="left" if c_idx in [2, 4, 6, 7, 8] else "center", vertical="center")
                c_cell.border = thin_border
            ws_extra.row_dimensions[2].height = 24

            for r_idx, er in enumerate(all_extra_records, 3):
                st = er.student
                sess = er.session
                crs = sess.course if sess else None
                fac_name = crs.teacher.full_name if (crs and crs.teacher) else (sess.teacher.full_name if (sess and sess.teacher) else "Faculty Coordinator")
                d_str = sess.session_date.isoformat() if (sess and sess.session_date) else "N/A"
                st_ctx = f"{st.program} Sem {st.semester} Div {st.section} ({st.department})" if st else "N/A"

                ws_extra.cell(row=r_idx, column=1, value=d_str).alignment = Alignment(horizontal="center", vertical="center")
                ws_extra.cell(row=r_idx, column=2, value=st.full_name if st else "Student").alignment = Alignment(horizontal="left", vertical="center")
                ws_extra.cell(row=r_idx, column=3, value=st.roll_number if st else "N/A").alignment = Alignment(horizontal="center", vertical="center")
                ws_extra.cell(row=r_idx, column=4, value=st_ctx).alignment = Alignment(horizontal="left", vertical="center")
                ws_extra.cell(row=r_idx, column=5, value=crs.code if crs else "N/A").alignment = Alignment(horizontal="center", vertical="center")
                ws_extra.cell(row=r_idx, column=6, value=crs.name if crs else (sess.session_name if sess else "N/A")).alignment = Alignment(horizontal="left", vertical="center")
                ws_extra.cell(row=r_idx, column=7, value=sess.session_name if sess else "N/A").alignment = Alignment(horizontal="left", vertical="center")
                ws_extra.cell(row=r_idx, column=8, value=fac_name).alignment = Alignment(horizontal="left", vertical="center")
                
                pres_cell = ws_extra.cell(row=r_idx, column=9, value=er.status)
                pres_cell.alignment = Alignment(horizontal="center", vertical="center")
                pres_cell.font = Font(name="Segoe UI", size=9, bold=True, color="15803D")
                pres_cell.fill = green_fill

                type_cell = ws_extra.cell(row=r_idx, column=10, value="EXTRA_LECTURE")
                type_cell.alignment = Alignment(horizontal="center", vertical="center")
                type_cell.font = Font(name="Segoe UI", size=9, bold=True, color="B45309")
                type_cell.fill = amber_light_fill

                for col in range(1, 11):
                    ws_extra.cell(row=r_idx, column=col).border = thin_border
                ws_extra.row_dimensions[r_idx].height = 19

            for col in range(1, 11):
                ws_extra.column_dimensions[get_column_letter(col)].width = 20

        filename = f"Attendance_Master_{uuid.uuid4().hex[:6]}.xlsx"
        filepath = str(settings.REPORTS_DIR / filename)
        wb.save(filepath)
        return filepath

    # =========================================================================
    # ADVANCED PDF EXPORT
    # =========================================================================
    @staticmethod
    def export_advanced_pdf(
        db: Session,
        class_id: Optional[int] = None,
        class_ids: Optional[Any] = None,
        department: Optional[str] = None,
        departments: Optional[Any] = None,
        program: Optional[str] = None,
        programs: Optional[Any] = None,
        semester: Optional[str] = None,
        semesters: Optional[Any] = None,
        division: Optional[str] = None,
        divisions: Optional[Any] = None,
        attendance_type: Optional[str] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> str:
        data = ReportService.get_advanced_report_data(
            db=db,
            class_id=class_id,
            class_ids=class_ids,
            department=department or departments,
            programs=programs or program,
            semesters=semesters or semester,
            divisions=divisions or division,
            attendance_type=attendance_type,
            start_date=start_date,
            end_date=end_date
        )

        filename = f"Attendance_Dossier_{uuid.uuid4().hex[:6]}.pdf"
        filepath = str(settings.REPORTS_DIR / filename)

        doc = SimpleDocTemplate(
            filepath,
            pagesize=landscape(A4),
            rightMargin=20,
            leftMargin=20,
            topMargin=20,
            bottomMargin=20
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "DocTitle",
            parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=13,
            leading=16,
            textColor=colors.HexColor("#0F172A"),
            alignment=1
        )
        sub_style = ParagraphStyle(
            "DocSub",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=11,
            textColor=colors.HexColor("#64748B"),
            alignment=1
        )
        batch_heading_style = ParagraphStyle(
            "BatchHead",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=9.5,
            leading=12,
            textColor=colors.HexColor("#1E3A8A"),
            spaceBefore=6,
            spaceAfter=3
        )
        table_hdr = ParagraphStyle(
            "THdr",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=7,
            leading=8.5,
            textColor=colors.white,
            alignment=1
        )
        cell_c = ParagraphStyle(
            "CC",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=6.5,
            leading=8,
            textColor=colors.HexColor("#1E293B"),
            alignment=1
        )
        cell_l = ParagraphStyle(
            "CL",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=6.5,
            leading=8,
            textColor=colors.HexColor("#1E293B"),
            alignment=0
        )

        elements = []
        elements.append(Paragraph("VisionAttend Pro — Comprehensive Attendance Dossier", title_style))
        elements.append(Paragraph(f"Academic Scope: {data['scope_title']} &bull; Date Range: {data['start_date'] or 'Start'} to {data['end_date'] or 'Present'} &bull; Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}", sub_style))
        elements.append(Spacer(1, 6))

        # Overview Table
        overview_data = [
            [
                Paragraph("<b>Total Sessions Held</b>", cell_c),
                Paragraph("<b>Total Students Enrolled</b>", cell_c),
                Paragraph("<b>Overall Average %</b>", cell_c),
                Paragraph("<b>Extra Lectures Conducted</b>", cell_c),
                Paragraph("<b>Defaulters Count (<75%)</b>", cell_c)
            ],
            [
                Paragraph(f"<b>{data['total_sessions_conducted']}</b>", cell_c),
                Paragraph(f"<b>{data['total_enrolled']}</b>", cell_c),
                Paragraph(f"<font color='#15803D'><b>{data['average_attendance_percentage']}%</b></font>", cell_c),
                Paragraph(f"<font color='#D97706'><b>{data.get('total_extra_lectures_conducted', 0)}</b></font>", cell_c),
                Paragraph(f"<font color='#B91C1C'><b>{data['defaulters_count']}</b></font>", cell_c)
            ]
        ]
        ov_t = Table(overview_data, colWidths=[1.8*inch, 1.8*inch, 1.8*inch, 1.8*inch, 1.8*inch])
        ov_t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
            ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#CBD5E1")),
            ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
            ("TOPPADDING", (0, 0), (-1, -1), 2.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5)
        ]))
        elements.append(ov_t)
        elements.append(Spacer(1, 8))

        # Batches
        batches = data.get("batches", [])
        for b in batches:
            elements.append(Paragraph(f"<b>{b['batch_label']}</b> — Enrolled: {b['total_enrolled']} | Sessions: {b['total_sessions_conducted']} | Avg: {b['average_attendance_percentage']}% | Extra: {b.get('extra_lectures_count', 0)} | Defaulters: {b['defaulters_count']}", batch_heading_style))
            
            b_rows = [
                [
                    Paragraph("<b>Roll No</b>", table_hdr),
                    Paragraph("<b>Student Full Name</b>", table_hdr),
                    Paragraph("<b>Norm. P</b>", table_hdr),
                    Paragraph("<b>Norm. A</b>", table_hdr),
                    Paragraph("<b>Norm. %</b>", table_hdr),
                    Paragraph("<b>Extra</b>", table_hdr),
                    Paragraph("<b>Tot. Sess</b>", table_hdr),
                    Paragraph("<b>Tot. P</b>", table_hdr),
                    Paragraph("<b>Final %</b>", table_hdr),
                    Paragraph("<b>Status</b>", table_hdr)
                ]
            ]
            t_style = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#94A3B8"))
            ]

            for idx, s in enumerate(b["students"], 1):
                is_def = s["is_defaulter"]
                s_col = "#B91C1C" if is_def else "#15803D"
                b_rows.append([
                    Paragraph(s["roll_number"], cell_c),
                    Paragraph(s["full_name"], cell_l),
                    Paragraph(str(s["normal_present"]), cell_c),
                    Paragraph(str(s["normal_absent"]), cell_c),
                    Paragraph(f"<b>{s['normal_percentage']}%</b>", cell_c),
                    Paragraph(f"<font color='#D97706'><b>{s['extra_lectures']}</b></font>", cell_c),
                    Paragraph(str(s["total_sessions"]), cell_c),
                    Paragraph(f"<font color='#15803D'><b>{s['total_present']}</b></font>", cell_c),
                    Paragraph(f"<font color='{s_col}'><b>{s['final_percentage']}%</b></font>", cell_c),
                    Paragraph(f"<font color='{s_col}'><b>{'DEFAULTER' if is_def else 'ELIGIBLE'}</b></font>", cell_c)
                ])
                if is_def:
                    t_style.append(("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#FEE2E2")))
                elif idx % 2 == 0:
                    t_style.append(("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#F8FAFC")))

            batch_table = Table(b_rows, colWidths=[1.0*inch, 2.2*inch, 0.65*inch, 0.65*inch, 0.75*inch, 0.65*inch, 0.75*inch, 0.65*inch, 0.85*inch, 0.95*inch])
            batch_table.setStyle(TableStyle(t_style))
            elements.append(batch_table)
            elements.append(Spacer(1, 6))

        doc.build(elements)
        return filepath

    @staticmethod
    def get_class_summary_data(db: Session, class_id: int, start_date: Optional[date] = None, end_date: Optional[date] = None) -> Dict[str, Any]:
        """
        Calculates high-level summary metrics for a specific class offering.
        """
        course = db.query(ClassCourse).filter(ClassCourse.id == class_id).first()
        if not course:
            return {}

        session_query = db.query(AttendanceSession).filter(AttendanceSession.class_id == class_id)
        if start_date:
            session_query = session_query.filter(AttendanceSession.session_date >= start_date)
        if end_date:
            session_query = session_query.filter(AttendanceSession.session_date <= end_date)

        sessions = session_query.all()
        session_ids = [s.id for s in sessions]
        total_sessions = len(sessions)

        students = list(course.students)
        total_enrolled = len(students)

        student_summaries = []
        total_pct = 0.0
        defaulters = 0

        for s in students:
            records = []
            if session_ids:
                records = db.query(AttendanceRecord).filter(
                    AttendanceRecord.session_id.in_(session_ids),
                    AttendanceRecord.student_id == s.id,
                    AttendanceRecord.is_extra_lecture == False,
                    AttendanceRecord.verification_type != "EXTRA_LECTURE",
                    AttendanceRecord.attendance_type != "EXTRA_LECTURE"
                ).all()
            
            present_count = sum(1 for r in records if r.status in ["PRESENT", "LATE"])
            absent_count = max(0, total_sessions - present_count)
            normal_pct = round((present_count / total_sessions) * 100.0, 2) if total_sessions > 0 else 0.0

            # Query extra lectures for student s
            extra_query = db.query(AttendanceRecord).join(AttendanceSession, AttendanceRecord.session_id == AttendanceSession.id).filter(
                AttendanceRecord.student_id == s.id,
                (AttendanceRecord.is_extra_lecture == True) | 
                (AttendanceRecord.verification_type == "EXTRA_LECTURE") | 
                (AttendanceRecord.attendance_type == "EXTRA_LECTURE")
            )
            if start_date:
                extra_query = extra_query.filter(AttendanceSession.session_date >= start_date)
            if end_date:
                extra_query = extra_query.filter(AttendanceSession.session_date <= end_date)
            extra_records = extra_query.all()
            extra_count = len(extra_records)

            total_student_sessions = total_sessions + extra_count
            total_student_present = present_count + extra_count
            total_student_absent = max(0, total_student_sessions - total_student_present)
            final_pct = round((total_student_present / total_student_sessions) * 100.0, 2) if total_student_sessions > 0 else 0.0
            is_defaulter = final_pct < settings.DEFAULTER_THRESHOLD_PERCENT if total_student_sessions > 0 else False

            if is_defaulter:
                defaulters += 1
            total_pct += final_pct

            student_summaries.append({
                "student_id": s.id,
                "roll_number": s.roll_number,
                "full_name": s.full_name,
                "normal_sessions": total_sessions,
                "normal_present": present_count,
                "normal_absent": absent_count,
                "normal_percentage": normal_pct,
                "extra_lectures": extra_count,
                "total_sessions": total_student_sessions,
                "total_present": total_student_present,
                "total_absent": total_student_absent,
                "final_percentage": final_pct,
                "percentage": final_pct,
                "attendance_percentage": final_pct,
                "present_sessions": total_student_present,
                "absent_sessions": total_student_absent,
                "is_defaulter": is_defaulter
            })

        avg_pct = round(total_pct / total_enrolled, 2) if total_enrolled > 0 else 0.0

        return {
            "class_id": course.id,
            "class_code": course.code,
            "class_name": course.name,
            "department": course.department,
            "program": getattr(course, "program", "B.Tech"),
            "semester": course.semester,
            "section": course.section,
            "teacher_name": course.teacher.full_name if course.teacher else "Unassigned",
            "total_sessions": total_sessions,
            "total_sessions_conducted": total_sessions,
            "total_enrolled": total_enrolled,
            "average_attendance_percentage": avg_pct,
            "defaulters_count": defaulters,
            "student_summaries": student_summaries,
            "students_summary": student_summaries
        }

    @staticmethod
    def export_excel(db: Session, class_id: Optional[int] = None, **kwargs) -> str:
        return ReportService.export_advanced_excel(db, class_id=class_id, **kwargs)

    @staticmethod
    def export_pdf(db: Session, class_id: Optional[int] = None, **kwargs) -> str:
        return ReportService.export_advanced_pdf(db, class_id=class_id, **kwargs)

    @staticmethod
    def get_all_defaulters(db: Session, class_id: Optional[int] = None) -> List[Dict[str, Any]]:
        """
        Scans all courses and returns students with attendance below threshold (<75%).
        """
        if class_id:
            courses = db.query(ClassCourse).filter(ClassCourse.id == class_id).all()
        else:
            courses = db.query(ClassCourse).all()
        defaulters = []
        for c in courses:
            summary = ReportService.get_class_summary_data(db, c.id)
            for s in summary.get("student_summaries", []):
                if s.get("is_defaulter"):
                    defaulters.append({
                        "student_id": s["student_id"],
                        "roll_number": s["roll_number"],
                        "student_name": s["full_name"],
                        "class_code": c.code,
                        "class_name": c.name,
                        "program": getattr(c, "program", "B.Tech"),
                        "semester": c.semester,
                        "section": c.section,
                        "attendance_percentage": s["percentage"]
                    })
        return defaulters

report_service = ReportService()
